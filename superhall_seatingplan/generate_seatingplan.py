from pathlib import Path
import matplotlib.pyplot as plt
import numpy as np
import sys,random,argparse
from MCR_Dining.superhall_seatingplan.setup import SetupMatrices
from MCR_Dining.getnames import AttendeeScraper

from MCR_Dining.superhall_seatingplan.metrics_moves import total_happiness,all_happiness,all_sat_with_guests,all_sat_with_friends
from MCR_Dining.superhall_seatingplan.cyth import sa_core
"""
============================================================
                 SEATING PLAN ALGORITHM
============================================================
Author: A. C. Hunt

Description:
------------
This module implements a seating plan optimization algorithm
using a combination of Python and Cython for performance. The
algorithm handles guest assignment, special constraints, and
preference scoring to generate reproducible, optimized seating
arrangements. It uses simulated annealing. 
"""

### Inputs
manual_removal=0 # Switch to do the manual removal of guests etc. described in xmas_superhall_fixes
verbose=0        # Switch to make the outputs more verbose

### file locations
folder='/home/ach221/Desktop'
event_booking_html = f"{folder}/Upay - Event Booking.html"
seating_form_responses = f"{folder}/Superhall_Seating_Request_Jan31"
swaps_xls = f"{folder}/MTSuperhallSwaps2025-26.xlsx"

### Set the seed with argparse
parser = argparse.ArgumentParser(description="Run seating / bath simulation with optional seed")
parser.add_argument("--seed", type=int, default=0, help="Random seed")
args = parser.parse_args()
np.random.seed(args.seed)
random.seed(args.seed)
sa_core.seed_c_rng(args.seed)
print(f'using seed {args.seed}')

### Get the names from Upay and seating form responses to generate the Matrices required
guestlist=AttendeeScraper(verbose,manual_removal)
guestlist.load_Upay(event_booking_html)
# guestlist.load_Swaps(swaps_xls)
guestlist.pretty_print()

### Get the Matrices for the propagation
MatMaker = SetupMatrices(guestlist,verbose,manual_removal)
# Specify the tables, their number of seats, and locations
if(0): # the whole hall
    table_types=['high','long','long','long','long']
    table_seats=[24,36,40,40,40]
    table_posns=np.array([[3,6],[8,6],[13,6],[18,6],[23,6]])

table_types=['high','long','long']
table_seats=[24,36,16]
MatMaker.specify_hall_params(table_types,table_seats,table_posns)

A,P,G,seat_positions,guestlist = MatMaker.get_Matrices(seating_form_responses)
ntot=A.shape[0]

### Randomize initial confign
s=np.arange(ntot,dtype=np.int32)
p=np.arange(ntot,dtype=np.int32)
s = np.random.permutation(ntot)
p = np.empty_like(s)
p[s] = np.arange(ntot)
h=total_happiness(A,P,G,p,s)

### Setup the plot
show=0
save_to_spreadsheet=0
if show:
    plt.ion()
    sc,cbar,ax,stop_button,text_labels=setupp.plot_setup(plt,seat_positions,all_happiness(A,P,G,p,s),p)
    def stop(event):sys.exit()
    stop_button.on_clicked(stop)

### Parameters   
T0 = 100
T = T0
hlist = []
nt = 2_000_000
nt = 2_0
all_hlist=[]
all_t=[]
cooling_rate = 0.99995
nhist = 50
tol = 0.1
h_best=0
p_best=p.copy()
h = total_happiness(A, P, G, p, s)


#convert everything to integers NOTE that this means that all weighting MUST be integers
s = s.astype(np.int32)
p = p.astype(np.int32)
s_trial=s.copy()
p_trial=p.copy()
def csr_to_int32(M): return M.indptr.astype(np.int32),M.indices.astype(np.int32),M.data.astype(np.int32)
A_indptr, A_indices, A_data = csr_to_int32(A)
P_indptr, P_indices, P_data = csr_to_int32(P)
G_indptr, G_indices, G_data = csr_to_int32(G)

valid_found=0
for it in range(nt):
    # monte carlo move
    delta_h,s_trial,p_trial,_ = sa_core.trial_move3(ntot, s,p,
                            A_indptr, A_indices, A_data,
                            P_indptr, P_indices, P_data,
                            G_indptr, G_indices, G_data)

    # Metropolis acceptance rule
    if delta_h > 0 or np.random.rand() < np.exp(delta_h / T):
        h += delta_h
        s[:] = s_trial
        p[:] = p_trial

    # help those who are pissed off once the time is late VERY AGGRESSIVE BIAS (+100)
    if it % 1000 == 0:
        # get the annoyed people
        score1,total1,pissed1=all_sat_with_guests(s,A,guestlist)
        outstr,npissed2,score2,total2,pissed2=all_sat_with_friends(s,A,P,guestlist)
        all_pissed=np.unique(np.concatenate([pissed1, pissed2]))
        # Output the situation
        print('SCORE1: {} of {}'.format(score1,total1))
        print(outstr)
        print(f'{it}/{nt}   h={h:.2f}   T={T:.3f}')
        # If noones mad, this is a valid solution
        if len(all_pissed) == 0:
            if h>h_best:
                h_best=h
                valid_found=1
                p_best=p.copy()
                s_best=s.copy() 
        #  Do forced moves of making the pissed people not mad:
        for pissed_indx in all_pissed:
            delta_h,s_trial,p_trial,bias = sa_core.trial_move3(ntot, s,p,
                            A_indptr, A_indices, A_data,
                            P_indptr, P_indices, P_data,
                            G_indptr, G_indices, G_data,int(pissed_indx))
            if delta_h+bias > 0 or np.random.rand() < np.exp((delta_h+bias) / T):
                h += delta_h
                s[:] = s_trial
                p[:] = p_trial
        hlist.append(h)

        if show:
            ax.set_title(f"Update {it+1}, happiness {h}")
            hall=all_happiness(A,P,G,p,s)
            sc.set_array(hall)  # update scatter color data
            for seat_number, t in enumerate(text_labels):
                t.set_text(p[seat_number])
            sc.set_clim(0, max(hall))       # update color scale range
            cbar.update_normal(sc)    # sync the colorbar with the new limits
            plt.draw()
            plt.pause(0.0001)

    # Check for local minimum (every 1000 steps maybe)
    if len(hlist) >= nhist and it % 1000 == 0:
        recent = np.array(hlist[-nhist:])
        mean = np.mean(recent)
        var = np.var(recent) / (mean**2 + 1e-9)
        if var < tol:
            print(f"Local minimum detected at iteration {it}. Reheating!")
            T *= 10   # reheat
            hlist = []  # reset history

    # Gradual cooling
    T *= cooling_rate
 
    # Append the h and the time to a list so we can plot later
    if it%(nt/1000)==0:
        all_hlist.append(int(h))  
        all_t.append(int(it))

# if we didnt find possible solutions, just use the lst one (we could just exit instead though to save memory)
if not valid_found:
    h_best=h
    p_best=p.copy()
    s_best=s.copy() 

## Save the results to the seating plan
p=p_best.copy()
s=s_best.copy()
import openpyxl
from openpyxl.styles import PatternFill
guest_fill = PatternFill(start_color="FFFF00", end_color="FFFF00",fill_type="solid")    # yellow
host_fill  = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")   # orange

print(f'best happiness {h_best}')

### Save the results to the seating plan xcell sheet
base = Path(__file__).parent # Get the path to the current script
filename = base / "Seating-plan-template.xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb.active 
# Write each name into its corresponding cell
for (col, row), person_indx in zip(seat_positions, p_best):
    name=guestlist.everyone[person_indx]
    cell=ws.cell(row=int(row), column=int(col), value=name)
    # add a fill  for the guests and the hosts 
    if name in guestlist.attendees_guest_map:
        if guestlist.attendees_guest_map[name]!=[]:
            cell.fill = host_fill
    if any(name in items for items in guestlist.attendees_guest_map.values()):
        cell.fill = guest_fill
# Save under a new name to keep the original template safe
wb.save(f"seating_filled.xlsx")

### Save the statistics for the fit 
score1,total1,_=all_sat_with_guests(s,A,guestlist)
outstr,npissed,score2,total2,_=all_sat_with_friends(s,A,P,guestlist)
h = total_happiness(A, P, G, p, s)
data = np.array([[score1, total1,score2,total2, npissed, h,args.seed]])
np.savetxt("results.txt", data, 
           header="# score1 total1 score2 total2 number_pissed_off total_hapiness seed", 
           fmt="%.4f", 
           comments="")

### save happiness graphs
data=np.array([all_t,all_hlist])
# np.savetxt("h.txt", data.T,header="nt h", comments="")
### plot the h
plt.xlabel('t')
plt.ylabel('h')
plt.plot(all_t,all_hlist)    
plt.title(f'Simulated annealing of seating-plan seed {args.seed}')   
plt.savefig('hplot.pdf')
# sys.exit()
if show:
    plt.ioff()  # turn off interactive mode when done
    plt.show()



