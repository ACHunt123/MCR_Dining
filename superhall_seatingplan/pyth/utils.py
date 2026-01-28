from matplotlib.widgets import Button
import openpyxl
from openpyxl.styles import PatternFill

def fill_spreadsheet(template,seat_positions,p_best,guestlist):
    guest_fill = PatternFill(start_color="FFFF00", end_color="FFFF00",fill_type="solid")    # yellow
    host_fill  = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")   # orange

    ### Save the results to the seating plan xcell sheet    
    wb = openpyxl.load_workbook(template)
    ws = wb.active 
    # Write each name into its corresponding cell
    for (col, row), person_indx in zip(seat_positions, p_best):
        name=guestlist.everyone[person_indx]
        cell=ws.cell(row=int(row), column=int(col), value=name)
        # add a fill  for the guests and the hosts 
        if name in guestlist.attendees_guest_map:
            if guestlist.attendees_guest_map[name]!=[]:
                cell.fill = host_fill
        if any(name in items for items in guestlist.attendees_guest_map.values()):
            cell.fill = guest_fill
    # Save under a new name to keep the original template safe
    outname = "seating_filled.xlsx"
    wb.save(outname)

def plot_setup(plt,seat_positions,happiness,p,mode='interactive'):
        ### Setup the plot
        if mode=='interactive':plt.ion()
        fig, ax = plt.subplots()
        ax.invert_yaxis() #for excel-like indexing
        sc = ax.scatter([x for x,y in seat_positions],[y for x,y in seat_positions],  c=happiness, cmap='RdYlGn')
        cbar=plt.colorbar(sc,label='n value')
        # Button setup
        button_ax = plt.axes([0.4, 0.05, 0.2, 0.075])  # x, y, width, height
        stop_button = Button(button_ax, 'STOP', color='lightcoral', hovercolor='red')
        text_labels = []
        for seat_number, (x, y) in enumerate(seat_positions):
            t = ax.text(x, y, p[seat_number], fontsize=9, ha='center', va='center', color='black')
            text_labels.append(t)
        return sc,cbar,ax,stop_button,text_labels
